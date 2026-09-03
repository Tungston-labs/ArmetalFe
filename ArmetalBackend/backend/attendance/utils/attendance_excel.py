from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_attendance_excel(
    employees,
    year,
    month,
    build_calendar_func
):
    """
    Generate complete attendance Excel for all employees.

    Returns:
        BytesIO object
    """

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Attendance"

    # =========================================================
    # STYLES
    # =========================================================

    thin_side = Side(
        style="thin",
        color="000000"
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    title_font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    section_font = Font(
        bold=True,
        size=12,
        color="FFFFFF"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    label_font = Font(
        bold=True
    )

    title_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    section_fill = PatternFill(
        fill_type="solid",
        fgColor="5B9BD5"
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="4472C4"
    )

    label_fill = PatternFill(
        fill_type="solid",
        fgColor="EAF2F8"
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    left_alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    # =========================================================
    # DATE RANGE
    # =========================================================

    import calendar
    from datetime import date

    start_date = date(
        year,
        month,
        1
    )

    last_day = calendar.monthrange(
        year,
        month
    )[1]

    end_date = date(
        year,
        month,
        last_day
    )

    month_name = start_date.strftime("%B %Y")

    # =========================================================
    # TABLE COLUMN COUNT (used for merged section/title widths)
    # =========================================================

    TABLE_COLUMN_COUNT = 10

    # =========================================================
    # ROW POINTER
    # =========================================================

    row = 1

    # =========================================================
    # EMPLOYEE LOOP
    # =========================================================

    for employee in employees:

        # -----------------------------------------------------
        # Build attendance calendar
        # -----------------------------------------------------

        (
            working_days,
            present_days,
            absent_days,
            lop_days,
            daily_records
        ) = build_calendar_func(
            employee,
            start_date,
            end_date
        )

        # =====================================================
        # TITLE
        # =====================================================

        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=TABLE_COLUMN_COUNT
        )

        cell = worksheet.cell(
            row=row,
            column=1
        )

        cell.value = (
            f"MONTHLY ATTENDANCE REPORT - "
            f"{month_name}"
        )

        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = center_alignment

        row += 2

        # =====================================================
        # EMPLOYEE DETAILS HEADING
        # =====================================================

        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=2
        )

        cell = worksheet.cell(
            row=row,
            column=1
        )

        cell.value = "Employee Details"
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = left_alignment

        row += 1

        # =====================================================
        # EMPLOYEE DETAILS
        # =====================================================

        details = [
            (
                "Employee Name",
                employee.name
            ),
            (
                "User Name",
                employee.employee_id
            ),
            (
                "Department",
                (
                    employee.department.name
                    if employee.department
                    else "-"
                )
            ),
            (
                "Working Days",
                working_days
            ),
            (
                "Present Days",
                present_days
            ),
            (
                "Absent Days",
                absent_days
            ),
            (
                "LOP Days",
                lop_days
            ),
        ]

        for label, value in details:

            label_cell = worksheet.cell(
                row=row,
                column=1
            )

            value_cell = worksheet.cell(
                row=row,
                column=2
            )

            label_cell.value = label
            value_cell.value = (
                value
                if value is not None
                else "-"
            )

            label_cell.font = label_font
            label_cell.fill = label_fill
            label_cell.border = border

            value_cell.border = border

            row += 1

        row += 1

       

        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=TABLE_COLUMN_COUNT
        )

        cell = worksheet.cell(
            row=row,
            column=1
        )

        cell.value = "Daily Attendance"
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = left_alignment

        row += 1

        # =====================================================
        # TABLE HEADER
        # =====================================================

        headers = [
            "Date",
            "Status",
            "Attendance Type",
            "Punch In",
            "Punch Out",
            "Total Hours",
            "Note",
            "Updated By",
            "Role",
            "Updated At",
        ]

        for column, header in enumerate(
            headers,
            start=1
        ):

            cell = worksheet.cell(
                row=row,
                column=column
            )

            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        row += 1

        # =====================================================
        # DAILY RECORDS
        # =====================================================

        for record in daily_records:

            attendance_type = record.get("attendance_type") or "-"

            values = [
                (
                    record["date"].strftime("%Y-%m-%d")
                    if record.get("date")
                    else "-"
                ),
                record.get(
                    "status"
                ) or "-",
                (
                    attendance_type.capitalize()
                    if attendance_type != "-"
                    else "-"
                ),
                record.get(
                    "first_punch_in"
                ) or "-",
                record.get(
                    "last_punch_out"
                ) or "-",
                (
                    record.get(
                        "total_hours"
                    )
                    if record.get(
                        "total_hours"
                    ) is not None
                    else "-"
                ),
                record.get(
                    "remark"
                ) or "-",
                record.get(
                    "updated_by"
                ) or "-",
                record.get(
                    "updated_by_role"
                ) or "-",
                record.get(
                    "updated_at"
                ) or "-",
            ]

            for column, value in enumerate(
                values,
                start=1
            ):

                cell = worksheet.cell(
                    row=row,
                    column=column
                )

                cell.value = value
                cell.alignment = center_alignment
                cell.border = border

            row += 1

        # Space between employees
        row += 2

    # =========================================================
    # COLUMN WIDTHS
    # =========================================================

    base_widths = {
        1: 16,   # Date
        2: 30,   # Status / Employee Detail Value (fits usernames/emails cleanly)
        3: 16,   # Attendance Type
        4: 42,   # Punch In (spacious for multiple comma-separated punch times)
        5: 42,   # Punch Out (spacious for multiple comma-separated punch times)
        6: 14,   # Total Hours
        7: 28,   # Note
        8: 20,   # Updated By
        9: 14,   # Role
        10: 24,  # Updated At ("DD MMM YYYY, HH:MM AM/PM")
    }

    # Calculate dynamic column width based on content length
    for col_idx in range(1, TABLE_COLUMN_COUNT + 1):
        col_letter = get_column_letter(col_idx)
        max_content_length = 0

        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            # Skip merged title and banner rows
            if cell.coordinate in worksheet.merged_cells:
                continue
            val = cell.value
            if val:
                val_str = str(val)
                if "MONTHLY ATTENDANCE REPORT" not in val_str and "Daily Attendance" not in val_str:
                    max_content_length = max(max_content_length, len(val_str))

        default_width = base_widths.get(col_idx, 16)
        worksheet.column_dimensions[col_letter].width = max(default_width, max_content_length + 3)

    # =========================================================
    # ROW HEIGHT
    # =========================================================

    for row_number in range(
        1,
        worksheet.max_row + 1
    ):

        worksheet.row_dimensions[
            row_number
        ].height = 22

    # =========================================================
    # SAVE TO MEMORY
    # =========================================================

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output